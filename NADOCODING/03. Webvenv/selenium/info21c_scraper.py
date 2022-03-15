import selenium
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from selenium import webdriver
from datetime import datetime 
from email import encoders
from random import randint
from time import sleep
import sys, os, errno, smtplib

# info21c.net ID / PWD / URL
ID = 'dyeng2019'
PWD = 'Sutra0615!'
URL = 'http://infose.info21c.net/info21c/bids/list/index?bidtype=con&bid_suc=bid&division=3'

# EMAIL ID / PWD
SENDER_ID = 'dydyenen@gmail.com' # 발신자
SENDER_PWD = '0314922449' # 발신자 비밀번호
RECIPIENTS = ['dydyenen@gmail.com'] # e.g. ['dydyenen@gmail.com', 'andy@daeyangeng.net', 'dyeng@daeyangeng.net']

# BSSAMT MIN, MAX
MAX_BSSAMT = 4500000000 # 기초 금액 최댓값
MIN_BSSAMT = 100000000 # 기초 금액 최솟값

# TODAY
TODAY = datetime.now().strftime('%m.%d')
YEAR = datetime.now().strftime('%Y')
LodDay = datetime.now().strftime('%Y-%m-%d [%H:%M:%S]')

# LOCATIONS - 해당 위치에 반드시 존재해야하며 수정 시 위치 변경 필수
RECENT_BIDPLINK_LOCATION = './recent_info21.txt' # 마지막 입찰 공고 링크를 담고있는 텍스트 파일
FORM_LOCATION = './form.xlsx' # 엑셀 양식 위치
LOG_LOCATION = './log.txt' # 로그 파일 위치
FILE_NAME = f'./output/인공지능을 이용한 투찰가 산정결과-{TODAY}.xlsx' # 오늘 날짜 폴더 위치
CHROME_LOCATION = './chromedriver.exe' # 크롬 드라이버 위치

# FORM XLSX SHEET NAME
SHEET_NAME = 'form'

class BidbplancScraper():
    def __init__(self):
        self.driver = None # Web Driver
        self.LOG = None # 로그 파일
        self.bidpblancLinks = [] # 입찰 공고 링크를 담을 리스트

    def signIn(self):
        # info21c.net 로그인
        userid = self.driver.find_element_by_id('userid')
        userid.clear()
        userid.send_keys(ID)
        
        userpw = self.driver.find_element_by_id('userpw')
        userpw.clear()
        userpw.send_keys(PWD)
        
        self.driver.find_element_by_xpath('//*[@id="btn_login"]/img').click()

    def setPageSize(self, size = '15'): # 스크롤 때문에 15로 고정 
        # 입찰 공고 목록 수 지정
        pageSize = self.driver.find_element_by_xpath('//*[@id="w0"]/div[1]/div[1]/select[1]') # //*[@id="w0"]/div[1]/div[1]/select[1]
        selector = Select(pageSize)
        selector.select_by_value(str(size)) 

    def getBidLink(self, startPage = 1, endPage = 7):  # 7 페이지까지 긁으면 약 한달 분량
        try:
            """ 
            [16324:8176:0429/131657.791:ERROR:browser_switcher_service.cc(238)] XXX Init()
            [16324:8176:0429/131704.435:ERROR:device_event_log_impl.cc(162)] [13:17:04.434] Bluetooth: bluetooth_adapter_winrt.cc:1055 Getting Default Adapter failed.
            위 로그가 출력되는 것을 방지하기 위해 아래 ChromeOptions 추가함
            """
            option = webdriver.ChromeOptions()
            option.add_experimental_option('excludeSwitches', ['enable-logging'])
            self.driver = webdriver.Chrome(CHROME_LOCATION, options = option) # Chrome Driver (chrome_options is deprecated)

            # web driver 호출 후 약간의 딜레이
            self.driver.implicitly_wait(3)

            # 주소로 이동
            self.driver.get(URL)

            # 로그인 및 페이지 설정
            self.signIn()
            self.setPageSize()

            for pageNo in range(startPage + 2, endPage + 3):  # 페이지가 시작이 1으므로 +2를 해줌 range(0, 6) = [0, 1, 2, 3, 4, 5]
                tbody = self.driver.find_element_by_xpath('//*[@id="w0"]/table/tbody') # //*[@id="w0"]/table
                rows = tbody.find_elements_by_css_selector('tr')

                # 테이블 행마다 링크를 가져옴
                for row in rows:
                    list_link = row.find_element_by_class_name('list_link') # //*[@id="w0"]/table/tbody/tr[1]/td[2]/a
                    if list_link:
                        self.bidpblancLinks.append(list_link.get_attribute('href'))
                        # print(len(self.bidpblancLinks))
                
                # 1 ~ 5초 무작위로 대기
                sleep(randint(1, 5))

                # 다음 목록으로 넘김
                self.driver.find_element_by_xpath(f'//*[@id="w0"]/div[3]/ul/li[{pageNo}]/a').click() # //*[@id="w0"]/div[3]/ul/li[3]/a

        except Exception as e:
            self.error(f'getBidLink() - Failed to get the bidpblanc link\n{e}\n')

    def getLastBidpblancLink(self):
        # 마지막 입찰 공고를 읽고 최신 입찰 공고로 업데이트 함
        # 경로에 있어야하며, 파일에 원하는 공고보다 먼저 나온 공고 링크를 넣어주어야 함
        with open(RECENT_BIDPLINK_LOCATION, 'r', encoding = 'utf-8') as f:
            lastBidpblancLink = f.readline() 
            if not(lastBidpblancLink):
                self.error(f'getLastBidpblancLink() - {RECENT_BIDPLINK_LOCATION} is empty\n')
        with open(RECENT_BIDPLINK_LOCATION, 'w', encoding = 'utf-8') as f:
            f.write(self.bidpblancLinks[0])

        return lastBidpblancLink
    
    def removeOldBidpblancLink(self, lastBidpblancLink):
        # 마지막으로 작업한 공고 링크를 참조해서 추출한 링크 중 새로운 링크만 남김
        # 마지막 입찰 공고 링크가 새로 추출한 링크에 존재하지 않는 경우 홈페이지에서 직접 확인 필요
        if lastBidpblancLink in self.bidpblancLinks:
            self.bidpblancLinks = self.bidpblancLinks[:self.bidpblancLinks.index(lastBidpblancLink)]
        else:
            self.LOG.write('removeOldBidpblanc() - Failed to find lastBidpblanc in bidpblancs\n')
            self.LOG.write(f'                       PLEASE CHECK {RECENT_BIDPLINK_LOCATION}\n')

    def getBidDetails(self):
        # 엑셀 불러오기
        form = load_workbook(FORM_LOCATION)
        sheet = form[SHEET_NAME]
        bidCount = 0
        last_row = sheet.max_row

        # 테두리 양식 설정
        border = Border(left = Side(style = 'thin', color = '000000'),
                        right = Side(style = 'thin', color = '000000'),
                        top = Side(style = 'thin', color = '000000'),
                        bottom = Side(style = 'thin', color = '000000'))


        # reversed()를 사용하여 먼저 공고된 순으로 처리
        for link in reversed(self.bidpblancLinks):
            try:
                # 무작위로 2 ~ 9초 딜레이
                sleep(randint(2, 5))

                self.driver.get(link)
                
                bidNtceNo = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[1]/td[2]').text # 공고번호-차수 //*[@id="basicInfo"]/table/tbody/tr[1]/td[2]
                bidNtceNm = self.driver.find_element_by_xpath('//*[@id="top_wrap"]/div[2]/table/tbody/tr/td').text # 공고명 //*[@id="top_wrap"]/div[2]/table/tbody/tr/td
                labels = self.driver.find_element_by_xpath('//*[@id="top_wrap"]/div[2]/table/tbody/tr/td').find_elements_by_css_selector('label')
                if labels: # 공고명 뒤에 붙는 수식어구 일부 제거
                    temp = ""
                    for label in labels:
                        temp = temp + label.text
                    
                    if "취소" in temp:
                        continue

                    bidNtceNm = bidNtceNm[:bidNtceNm.rfind(temp)]
                
                ntceInsttNm = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[3]/td[2]').text.replace('발주처심층분석', '').replace('\n', '') # 발주기관 
                sucsfbidLwltRate = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[6]/td[2]').text # 낙찰하한율
                ntceInsttOfcl = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[5]/td[2]').text # 담당자 이름 / 전화번호
                # dminsttNm = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[4]/td[1]').text # 수요처
                # rsrvtnPrceRngRate = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[6]/td[1]').text # 예가 변동폭

                hasAValue = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[11]/th').text == 'A값'
              
                if hasAValue: # A값 상세금액 항목이 존재하는 경우 표가 1칸 씩 밀림
                    bidQlfctRgstDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[8]/td[1]').text  # 참가등록마감
                    bidBeginDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[8]/td[2]').text  # 입찰개시일
                    bidClseDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[9]/td[1]/span').text  # 투찰마감일시
                    
                    opengDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[10]/td[1]').text  # 입찰일시
                    hasOCost = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[12]/th').text == '순공사원가'
                    if hasOCost:
                        bssamt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[13]/td[1]/b').text # 기초금액
                    else:
                        bssamt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[12]/td[1]/b').text  # 기초금액
                    aValue = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[11]/td[1]/span').text # A 값
                    try: # e.g. 궁평항 건설 전기공사 A값 적용이지만 값이 없는 경우 예외 처리 
                        aValue = int(aValue.split()[3].replace(',', '')) 
                    except:
                        aValue = ""
                else:
                    bidQlfctRgstDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[8]/td[1]').text # 참가등록마감
                    bidBeginDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[8]/td[2]').text # 입찰개시일
                    bidClseDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[9]/td[1]/span').text # 투찰마감일시
                    opengDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[10]/td[1]').text  # 입찰일시
                    hasOCost = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[11]/th').text == '순공사원가'
                    if hasOCost:
                        bssamt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[12]/td[1]/b').text # 기초금액
                    else:
                        bssamt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[11]/td[1]/b').text  # 기초금액
                    opengDt = self.driver.find_element_by_xpath('//*[@id="basicInfo"]/table/tbody/tr[10]/td[1]').text # 입찰일시
                
                try: 
                    # 기초금액을 제공하지 없는 경우, 기초금액이 최댓값을 초과하거나 최솟값 미만인 경우 제외
                    bssamt = int(bssamt.replace(' 원', '').replace(',', ''))
                    if bssamt > MAX_BSSAMT or bssamt < MIN_BSSAMT:
                        raise ValueError
                except Exception as e:
                    continue
                
                # 낙찰하한율이 없는 경우 제외
                if sucsfbidLwltRate == '': 
                    continue
                
                # 번호, 공고번호, 공고명, 발주기관, 담당자, 참가등록마감, 입찰개시일, 투찰마감일시, 개찰일시, 낙찰하한율, 기초금액, 투찰가, 투찰하기링크
                basicG2bLink = 'http://www.g2b.go.kr/pt/menu/selectSubFrame.do?framesrc=/pt/menu/frameTgong.do?url='
                information = [bidCount + 1, bidNtceNo, bidNtceNm, ntceInsttNm, ntceInsttOfcl, bidQlfctRgstDt, bidBeginDt, bidClseDt, opengDt, float(sucsfbidLwltRate.rstrip('%')) / 100, bssamt, '', '투찰하기']

                for col in range(1, len(information) + 1):                   
                    if col == 6 or col == 9: # 참가등록마감, 개찰일시 빨간색 폰트 설정
                        sheet.cell(row = last_row, column = col).font = Font(color = 'FF0000')
                    elif col == 10: # 낙찰하한율 출력 형식 설정 e.g. 87.765%
                        sheet.cell(row = last_row, column = col).font = Font(color = 'FF0000')
                        sheet.cell(row = last_row, column = col).number_format = '00.000%'
                    elif col == 11 or col == 12: # 기초금액 및 투찰가 출력 형식 설정 e.g. 100,000
                        sheet.cell(row = last_row, column = col).number_format = '0,000'
                    
                    sheet.cell(row = last_row, column = col).value = information[col - 1]

                    if col == 13: # 투찰하기 링크 설정
                        try:
                            try: # 차수가 없는 경우 번호로만 사용
                                bidno, bidseq = bidNtceNo.split('-')
                            except:
                                bidno = bidNtceNo

                            # 검색
                            g2bLink = f'http://www.g2b.go.kr:8101/ep/tbid/tbidList.do?bidSearchType=1&bidno={bidno}&searchType=2'
                            
                            self.driver.get(g2bLink)
                       
                            # 있는 경우 하이퍼링크 처리 2개 이상 발견되는 경우가 있어서 -을 더해줌
                            g2bDetailLink = self.driver.find_element_by_partial_link_text(bidno + "-").get_attribute('href')
                 
                            sheet.cell(row = last_row, column = col).hyperlink = basicG2bLink + g2bDetailLink
                            sheet.cell(row = last_row, column = col).style = "Hyperlink"

                        except: # 검색결과 없는 경우 
                            sheet.cell(row = last_row, column = col).value = '' 
                                
                    sheet.cell(row = last_row, column = col).border = border
                    
                    if col == 13: # 텍스트 정렬
                        sheet.cell(row = last_row, column = col).alignment = Alignment(wrap_text = True, vertical = 'center', horizontal = 'center') # 자동 줄바꿈으로 자동 개행을 원하지 않을 경우 주석
                    else:
                        sheet.cell(row = last_row, column = col).alignment = Alignment(wrap_text = True, vertical = 'center') # 자동 줄바꿈으로 자동 개행을 원하지 않을 경우 주석

                bidCount += 1; last_row += 1

                # 파일명 제약 특수문자 제거
                # bidNtceNm = bidNtceNm.replace('/', "").replace('\\', "").replace(':', "").replace('?', "").replace('*', "").replace('<', "").replace('>', "").replace('', "").replace('"', "")
                
            except Exception as e:
                # 예외 발생 시 info21c.net의 입찰공고 상세 사이트의 xpath의 위치가 변경되어 값을 못 찾아오는 경우
                # 해당 항목을 새로운 위치로 변경해주어야 함
                self.error(f'getBidDetails() - Failed to get details({bidNtceNm})\n{e}\n')
        
        # 부합한 공고가 1개 이상인 경우 지정된 경로에 파일 저장
        if bidCount > 0:
            self.LOG.write(f'   {len(self.bidpblancLinks)}개 중 {bidCount}개 입찰 공고를 완료했습니다.\n')
            form.save(FILE_NAME)
        else:
            self.LOG.write(f'   {len(self.bidpblancLinks)}개 중 조건에 부합하는 입찰 공고가 존재하지 않습니다.\n')
        
        self.LOG.write('   Completed\n')
        self.LOG.close()
        self.driver.quit()

    def run(self):
        self.LOG = open(LOG_LOCATION, 'a')
        self.LOG.write(f'@{LodDay}\n')

        # 전기경기탭의 1페이지에서 7페이지까지 링크를 가져옴(15개씩)
        self.getBidLink()

        # 마지막 공고 번호 호출
        lastBidpblancLink = self.getLastBidpblancLink()

        # 최신 공고가 아닌 경우 제거
        self.removeOldBidpblancLink(lastBidpblancLink)
        if len(self.bidpblancLinks) == 0:
            # 제거 후 링크가 없는 경우, 종료
            self.error(f'There are no new bidpblancs {LodDay}\n')

        self.getBidDetails()

    def error(self, msg):
        # 오류 발생 시 오류 내용을 기록하고 드라이버 및 로그 파일을 종료 후 프로그램 종료
        self.LOG.write(msg)
        self.LOG.close()
        self.driver.quit()
        sys.exit(0)

if __name__ == "__main__":
    # MAIN
    bidpblancScraper = BidbplancScraper()
    bidpblancScraper.run()