from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져온다
# for row in ws.values:
#     for cell in row:
#         print(cell)

# 수식이 아닌 실제 데이터
# 계산되지 않은 데이터는 None
# 열었다 다시 저장하면 데이터 출력 가능
wb = load_workbook("sample_formula.xlsx", data_only= True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)