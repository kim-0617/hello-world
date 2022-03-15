from openpyxl import Workbook

wb = Workbook() # 새 워크북 생성 (excel 파일)
ws = wb.create_sheet() # 새로운 sheet 생성
ws.title = "Mysheet"
ws.sheet_properties.tabColor = "ff6699" # RGB형식으로 값을 넣어주면 색깔이 바뀜

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 sheet에 접근

print(wb.sheetnames) # 모든 sheet 이름 확인

new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
wb.close()
