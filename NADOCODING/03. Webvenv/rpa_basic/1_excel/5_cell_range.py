from cgi import print_directory
from openpyxl import Workbook, workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])

for i in range(1, 11):
    ws.append([i,randint(0,100),randint(0,100)])

# col_B = ws["B"] # 영어점수만 가져오기
# print(col_B) # 셀정보
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] # 영어, 수학 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# from openpyxl.utils.cell import coordinate_from_string # 셀의 좌표정보

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end =" ")
#         xy = coordinate_from_string(cell.coordinate) # 열과 행을 끊어서 튜플형식으로 보여줌
#         # print(xy, end = " ")
#         print(xy[0],end= " ") 
#         print(xy[1],end= " ") 
#     print()

# 전체 rows
# print(tuple(ws.rows)) # 한 줄씩 묶어서 튜플로 가져옴
# for row in tuple(ws.rows):
#     print(row) # A1 B1 C1 / A2 B2 C2 / ...
    # print(row[1].value)

# # 전체 columns
# print(tuple(ws.columns)) # 한 열씩 묶어서 튜플로 가져옴
# for column in tuple(ws.columns):
#     # print(column) # A 1~10 / B 1~10 / ...
#     print(column[1].value)

# for row in ws.iter_rows(): # 전체 row
#     print(row[1].value)

# for column in ws.iter_cols(): # 전체 col
#     print(column[0].value)

for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): # 2~11번줄 / 2~3열 만
    print(row[0].value)
    print(row[1].value)
    print("-"*50) # 모든학생의 영어,수학점수만 출력


wb.save("sample.xlsx")
wb.close()