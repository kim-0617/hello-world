from fractions import Fraction
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

ws.column_dimensions["A"].width = 5 # 너비
ws.row_dimensions[1].height = 50  # 높이 (1행의 높이)

a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색상은 빨갛게, 이탤릭, 두껍게
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column == 1: # A 번호열 제외
            continue

        # cell이 정수형 데이터이고 90점 이상이면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")

# 틀고정
ws.freeze_panes = "B2" # B2기준으로 틀 고정
# 좌우 스크롤해도 번호 나오고 / 위아래 스크롤해도 컬럼명 나오고

wb.save("sample_style.xlsx")