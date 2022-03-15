from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # 셀 객체정보
print(ws["A1"].value) # 값이 없을 땐 None 출력

print(ws.cell(row=1, column=1).value) # 1행 1열의 셀정보 (A1) 

c = ws.cell(row=1, column=3, value=10) # ws["C1"] = 10
print(c.value)

from random import Random, randint

# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1,11):
    for y in range(1,11):
        ws.cell(row = x, column = y, value = randint(1,100))

wb.save("sample.xlsx")
wb.close()