from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

# 데이터 삽입
ws.append(['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트', '총점', '성적'])
ws.append([1, 10, 8, 5, 14, 26, 12])
ws.append([2, 7, 3, 7, 15, 24, 18])
ws.append([3, 9, 5, 8, 8, 12, 4])
ws.append([4, 7, 8, 7, 17, 21, 18])
ws.append([5, 7, 8, 7, 16, 25, 15])
ws.append([6, 3, 5, 8, 8, 17, 0])
ws.append([7, 4, 9, 10, 16, 27, 18])
ws.append([8, 6, 6, 6, 15, 19, 17])
ws.append([9, 10, 10, 10, 19, 30, 19])
ws.append([10, 9, 8, 8, 20, 25, 20])

# 모든행 가운데 정렬
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 퀴즈2 오류로인한 만점처리
for row in ws.iter_rows(min_row=2):
    row[3].value = 10

total_row = [] # 현재 돌고있는 행의 총점
total = [] # 모든 행의 각 총점 정보
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=2, max_col=7): 
    for i in range(0,6):
        total_row.append(row[i].value)
    total.append(sum(total_row))
    total_row.clear()   

# 총점 삽입
for i in range(0,len(total)):
    ws.cell(i+2,8,total[i])

# 성적 삽입
for row in ws.iter_rows(min_col=2,min_row=2):
    if row[0].value < 5:
        row[7].value = 'F'
    elif row[6].value >= 90:
        row[7].value = 'A'
    elif row[6].value >= 80:
        row[7].value = 'B'
    elif row[6].value >= 70:
        row[7].value = 'C'
    else:
        row[7].value = 'D'
wb.save('scores.xlsx')